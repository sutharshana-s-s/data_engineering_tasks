import pandas as pd
from openpyxl import load_workbook

# Correct file mapping
SOURCE_FILES = {
    "apr": "MIS_Final_April.xlsx",
    "may": "MIS_Final_May.xlsx",
    "jun": "MIS_Final_June.xlsx"
}

OUTPUT_FILE = "Delta3_Final_Output.xlsx"


def extract_revenue_mapping(file_path):
    df = pd.read_excel(file_path, header=1)

    # Find Revenue row
    revenue_row = df[df.iloc[:, 0].astype(str).str.contains("Revenue", case=False, na=False)]

    if revenue_row.empty:
        return {}

    revenue_row = revenue_row.iloc[0]

    # Project names start from column index 2
    project_names = df.columns[2:]

    revenue_map = {}

    for col in project_names:
        project_name = str(col).strip().lower()
        value = revenue_row[col]

        if pd.notna(value):
            revenue_map[project_name] = value

    return revenue_map


# Load output file
wb = load_workbook(OUTPUT_FILE)
ws = wb.active


# Output column positions
MONTH_COLUMNS = {
    "apr": "D",   
    "may": "G",   
    "jun": "J"    
}


for month, file in SOURCE_FILES.items():
    revenue_map = extract_revenue_mapping(file)
    col_letter = MONTH_COLUMNS[month]

    for row in range(20, ws.max_row + 1):
        project_name = ws[f"A{row}"].value

        if not project_name:
            continue

        key = str(project_name).strip().lower()

        if key in revenue_map:
            ws[f"{col_letter}{row}"] = revenue_map[key]


wb.save("Delta3_Final_Output.xlsx")
