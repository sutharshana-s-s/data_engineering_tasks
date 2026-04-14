import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, date
import traceback

SOURCE_FILE = "Delta3_Apr.xlsx"
TEMPLATE_FILE = "Delta3_Output.xlsx"

# Extracts a month token from a value and returns a month string or None.
def get_month_token(value):
    if isinstance(value, (datetime, date)):
        token = value.strftime('%b').lower()
    elif not value:
        token = None
    else:
        token = str(value).strip().lower()[:3]
    return token

# Converts a string value to a float number.
def to_number(value):
    if isinstance(value, str):
        value = value.replace("£", "").replace(",", "").strip()
    return float(value)

if __name__ == "__main__":
    try:
        print("Starting Task 1 Excel processing")

        source_excel = pd.ExcelFile(SOURCE_FILE)
        wb = load_workbook(TEMPLATE_FILE)
        sheet = wb.active

        month_to_col = {}

        for col in range(4, 16):
            header = sheet.cell(row=6, column=col).value
            token = get_month_token(header)

            if token:
                month_to_col[token] = col

        for sheet_name in source_excel.sheet_names:
            sheet_token = get_month_token(sheet_name)

            if not sheet_token or sheet_token not in month_to_col:
                print(f"Skipping sheet: {sheet_name}")
                continue

            col = month_to_col[sheet_token]

            df = pd.read_excel(source_excel, sheet_name=sheet_name, header=None).dropna(how='all')
            data = df.set_index(0)[1].to_dict()

            revenue = to_number(data.get("Revenue", 0))

            total_cost = (
                to_number(data.get("Salary - Core employees", 0)) +
                to_number(data.get("Salary - TL / Managers", 0)) +
                to_number(data.get("Salary - Consultants", 0)) +
                to_number(data.get("Performance payments - Incentive & Others", 0))
            )

            sheet.cell(row=7, column=col, value=revenue)
            sheet.cell(row=8, column=col, value=data.get("Revenue %", ""))

            sheet.cell(row=10, column=col, value=total_cost)
            sheet.cell(row=11, column=col, value=total_cost / revenue if revenue else 0)

            sheet.cell(row=13, column=col, value=data.get("Total salary allocation for project", ""))
            sheet.cell(row=14, column=col, value=data.get("Total salary allocation %", ""))

        wb.save("Delta3_Output.xlsx")
        print("Task 1 Excel processing completed successfully.")
    except Exception as e:
        print(f"Error in Task 1 Excel processing: {e}")
        traceback.print_exc()